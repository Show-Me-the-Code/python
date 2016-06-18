#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''第 0014 题： 纯文本文件 student.txt为学生信息, 里面的内容（包括花括号）如下所示：
{
    "1":["张三",150,120,100],
    "2":["李四",90,99,95],
    "3":["王五",60,66,68]
}
请将上述内容写到 student.xls 文件中。'''

__author__ = 'Drake-Z'

import json
from collections import OrderedDict
from openpyxl import Workbook

def txt_to_xlsx(filename):
    file = open(filename, 'r', encoding = 'UTF-8')
    file_cintent = json.load(file, encoding = 'UTF-8')
    print(file_cintent)
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    for i in range(1, len(file_cintent)+1):
        worksheet.cell(row = i, column = 1).value = i
        for m in range(0, len(file_cintent[str(i)])):
            worksheet.cell(row = i, column = m+2).value = file_cintent[str(i)][m]
    workbook.save(filename = 'student.xls')

if __name__ == '__main__':
    txt_to_xlsx('student.txt')
