#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''第 0015 题： 纯文本文件 city.txt为城市信息, 里面的内容（包括花括号）如下所示：
{
    "1" : "上海",
    "2" : "北京",
    "3" : "成都"
}
请将上述内容写到 city.xls 文件中。'''

__author__ = 'Drake-Z'

import json
from collections import OrderedDict
from openpyxl import Workbook

def txt_to_xlsx(filename):
    file = open(filename, 'r', encoding = 'UTF-8')
    file_cintent = json.load(file, encoding = 'UTF-8')
    print(file_cintent)
    workbook = Workbook()
    worksheet = workbook.worksheets[0]
    for i in range(1, len(file_cintent)+1):
        worksheet.cell(row = i, column = 1).value = i
        worksheet.cell(row = i, column = 2).value = file_cintent[str(i)]
    workbook.save(filename = 'city.xls')

if __name__ == '__main__':
    txt_to_xlsx('city.txt')