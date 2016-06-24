from openpyxl import Workbook
from openpyxl import load_workbook

path_save = "/home/lwh/SublimeTextProject/test.xlsx"
path_read = "/home/lwh/SublimeTextProject/demo.txt"


def write_test():
    wb = Workbook()     # 1.creat a work book
    wb_sheet0 = wb.create_sheet("student")  # 2.create a sheet in the work book

    datas_str = ""
    datas_dict = {}

    with open(path_read, "r") as f:
        lines = f.readlines()

    # process the data from txt file.I make it be a dict
    for line in lines:
        if len(line.strip()) > 1:
            key, value = line.split(':')
            key = key.strip().strip("\"")
            value = value.strip().strip(",").strip("[").strip("]")
            value_list = value.split(",")
            # print(len(value_list))
            datas_dict[key] = value_list

    for i in range(1, 4):       # 3.assign value to cell of the sheet
        for j in range(1, 5):
            cell = wb_sheet0.cell(row=i, column=j + 1)
            cell.value = datas_dict[str(i)][j - 1]
            # print(cell.value)
        cell = wb_sheet0.cell(row=i, column=1)
        cell.value = i
    '''
    for row in wb_sheet0.iter_rows():
        for cell in row:
            print(cell.value)
    '''
    wb.save(path_save)      # 4.save the file
    print('success')

write_test()
