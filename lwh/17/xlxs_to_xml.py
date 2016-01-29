from xml.etree.ElementTree import ElementTree, Element, SubElement, Comment
from openpyxl import load_workbook
import json
from pprint import pprint


class ExcelToXml(object):

    def __init__(self):
        self.path_save = '/home/lwh/SublimeTextProject/test.xml'
        self.path_read = '/home/lwh/SublimeTextProject/test.xlsx'
        self.excel_dict = self.get_excel_dict()
        self.data = self.get_text(self.excel_dict)
        self.xml_final = self.get_xml(self.data)
        self.save_xml(self.xml_final)

    def get_excel_dict(self):
        wb = load_workbook(self.path_read)
        wb_sheet_student = wb['student']
        excel_dict = {}
        for row in wb_sheet_student.iter_rows():
            key = str(row[0].value).strip()
            value_list = row[1:]
            value = [ele.value for ele in value_list]
            excel_dict[key] = str(value)
        return excel_dict

    def get_text(self, excel_dict):
        res = '{\r\n'
        for key, val in excel_dict.items():
            res += ' ' + str(key) + ':' + val + '\n'
        res += '}\n'
        # print(res)
        return res

    def get_xml(self, data):
        root = Element('root')
        student = SubElement(root, 'students')
        student.append(Comment('学生信息表 "id" : [名字, 数学, 语文, 英文]'))
        # student.
        print(data)
        student.text = data
        xml_final = ElementTree(root)
        # print(xml_final)
        return xml_final

    def save_xml(self, xml_to_save):
        xml_to_save.write(
            self.path_save, encoding="us-ascii", default_namespace=None, short_empty_elements=True)


if __name__ == '__main__':
    excel_to_xml_obj = ExcelToXml()
